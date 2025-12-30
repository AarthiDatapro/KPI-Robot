from flask import Flask, render_template, request, send_file, jsonify
import openpyxl
import random
from io import BytesIO
import os
import logging
import sys
import gc
from datetime import datetime

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('app.log')
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Configuration
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['MEMORY_BUFFER_SIZE'] = 1024 * 1024 * 50  # 50MB buffer size

# ---------------- FIXED PROBABILITIES ----------------
V_PROBABILITY = 0.90
MARKS_PROBABILITY = 0.85
ANSWER_PROBABILITY = 0.90

# ---------------- FIXED PROJECTS ----------------
projects = {
    1: "Exploratory Analysis of Iris Flower Measurements",
    2: "Chemical Composition Analysis of Italian Wines",
    3: "Descriptive Statistics of Tumor Cell Nuclei Characteristics",
    4: "Profiling Health Indicators in Diabetes Patients",
    5: "Physical Exercise and Physiological Measurements: A Correlation Study",
    6: "Socioeconomic and Housing Trends in 1990 California Census Data",
    7: "Demographic and Survival Trends of Titanic Passengers",
    8: "Restaurant Billing and Tipping Behavior Analysis",
    9: "Demographic and Economic Profile of US Census Data (1994)",
    10: "Bank Loan Applicant Profiles of Credit Data"
}

def log_memory_usage(prefix=""):
    """Log current memory usage"""
    import psutil
    process = psutil.Process(os.getpid())
    mem_info = process.memory_info()
    logger.info(f"{prefix}Memory usage: {mem_info.rss / (1024 * 1024):.2f}MB")

def cleanup_resources(*resources):
    """Explicitly clean up resources"""
    for resource in resources:
        if hasattr(resource, 'close'):
            try:
                resource.close()
            except Exception as e:
                logger.error(f"Error closing resource: {e}")
    gc.collect()

# ---------------- HELPERS ----------------
def col_letter_to_index(letter):
    return ord(letter.upper()) - 64

def parse_columns(text):
    return [col_letter_to_index(c.strip()) for c in text.split(",") if c.strip()]

def parse_rows(text):
    rows = set()
    for part in text.split(","):
        part = part.strip()
        if "-" in part:
            start, end = part.split("-")
            rows.update(range(int(start), int(end) + 1))
        else:
            rows.add(int(part))
    return sorted(rows)

def process_option_1(source_wb, dest_wb, sheets, columns, rows):
    """Process option 1: Mark 'v' in specified cells"""
    for sheet_name in sheets:
        ws_src = source_wb[sheet_name]
        ws_dest = dest_wb.create_sheet(title=sheet_name)
        
        # Copy all rows
        for row in ws_src.iter_rows(values_only=True):
            ws_dest.append(row)
            
        # Mark cells
        for col in columns:
            for row_idx in rows:
                if random.random() < V_PROBABILITY:
                    ws_dest.cell(row=row_idx, column=col).value = "v"

def process_option_2(source_wb, dest_wb, sheets, columns, rows, marks):
    """Process option 2: Add marks, team, and project info"""
    sheet_index = 0
    team_number = 1

    while sheet_index < len(sheets) and team_number <= 10:
        for _ in range(5):  # 5 members per team
            if sheet_index >= len(sheets):
                break

            sheet_name = sheets[sheet_index]
            ws_src = source_wb[sheet_name]
            ws_dest = dest_wb.create_sheet(title=sheet_name)
            
            # Copy all rows
            for row in ws_src.iter_rows(values_only=True):
                ws_dest.append(row)
                
            # Update team and project info
            ws_dest["D4"] = team_number
            ws_dest["D5"] = projects[team_number]

            # Update marks
            for col in columns:
                for row_idx in rows:
                    if random.random() < MARKS_PROBABILITY:
                        ws_dest.cell(row=row_idx, column=col).value = marks
                    else:
                        ws_dest.cell(row=row_idx, column=col).value = random.randint(
                            marks - 3, marks - 1
                        )

            sheet_index += 1
        team_number += 1

def process_option_3(source_wb, dest_wb, sheets, columns, rows, ref_col):
    """Process option 3: Generate answers based on reference column"""
    for sheet_name in sheets:
        ws_src = source_wb[sheet_name]
        ws_dest = dest_wb.create_sheet(title=sheet_name)
        
        # Copy all rows
        for row in ws_src.iter_rows(values_only=True):
            ws_dest.append(row)
            
        # Process answers
        for col in columns:
            for row_idx in rows:
                correct = ws_dest.cell(row=row_idx, column=ref_col).value
                if correct not in [1, 2, 3, 4]:
                    continue

                if random.random() < ANSWER_PROBABILITY:
                    ws_dest.cell(row=row_idx, column=col).value = correct
                else:
                    wrong = [x for x in [1, 2, 3, 4] if x != correct]
                    ws_dest.cell(row=row_idx, column=col).value = random.choice(wrong)

# ---------------- ROUTE ----------------
@app.route("/", methods=["GET", "POST"])
def index():
    start_time = datetime.now()
    logger.info("New request received")
    log_memory_usage("Before processing: ")

    if request.method == "POST":
        try:
            if 'file' not in request.files:
                logger.error("No file part in the request")
                return jsonify({"error": "No file part"}), 400

            file = request.files['file']
            if file.filename == '':
                logger.error("No selected file")
                return jsonify({"error": "No selected file"}), 400

            logger.info(f"Processing file: {file.filename}")
            
            # Always use read-only mode for better memory efficiency
            try:
                wb = openpyxl.load_workbook(file, read_only=True)
                logger.info("Using read-only mode for better memory efficiency")
            except Exception as e:
                logger.error(f"Error loading workbook: {str(e)}")
                return jsonify({"error": "Error processing the Excel file. It might be corrupted or too large."}), 400

            option = request.form.get("option")
            columns = parse_columns(request.form.get("columns", ""))
            rows = parse_rows(request.form.get("row_ranges", ""))

            if not all([option, columns, rows]):
                logger.error("Missing required parameters")
                return jsonify({"error": "Missing required parameters"}), 400

            sheets = wb.sheetnames[1:]  # skip first sheet
            logger.info(f"Processing {len(sheets)} sheets")

            # Create a new workbook for output
            output_wb = openpyxl.Workbook()
            
            # Remove default sheet created by openpyxl
            output_wb.remove(output_wb.active)
            
            # Process based on option
            try:
                if option == "1":
                    process_option_1(wb, output_wb, sheets, columns, rows)
                elif option == "2":
                    marks = int(request.form.get("marks", 0))
                    process_option_2(wb, output_wb, sheets, columns, rows, marks)
                elif option == "3":
                    ref_column = request.form.get("ref_column", "").strip()
                    if not ref_column:
                        return jsonify({"error": "Reference column is required for Option 3"}), 400
                    ref_col = col_letter_to_index(ref_column)
                    process_option_3(wb, output_wb, sheets, columns, rows, ref_col)
                else:
                    return jsonify({"error": "Invalid option"}), 400

                # Save to memory-efficient buffer
                output = BytesIO()
                logger.info("Saving workbook...")
                output_wb.save(output)
                output.seek(0)
                
                # Clean up
                cleanup_resources(wb, output_wb, file)
                log_memory_usage("After processing: ")

                logger.info(f"Request processed in {(datetime.now() - start_time).total_seconds():.2f} seconds")
                
                return send_file(
                    output,
                    as_attachment=True,
                    download_name="updated.xlsx",
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                logger.error(f"Error during processing: {str(e)}", exc_info=True)
                if 'output_wb' in locals():
                    cleanup_resources(wb, output_wb, file)
                else:
                    cleanup_resources(wb, file)
                return jsonify({"error": "An error occurred while processing your request"}), 500

        except Exception as e:
            logger.error(f"Error processing request: {str(e)}", exc_info=True)
            return jsonify({"error": "An error occurred while processing your request"}), 500

    return render_template("index.html")

if __name__ == "__main__":
    port = int(os.environ.get('PORT', 10000))
    logger.info(f"Starting server on port {port}")
    app.run(host='0.0.0.0', port=port)
